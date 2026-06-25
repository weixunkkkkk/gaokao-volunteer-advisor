#!/usr/bin/env python3
"""Generate Gaokao school and major recommendations from normalized CSV data."""

from __future__ import annotations

import argparse
import csv
import json
import re
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATA_DIR = ROOT / "assets" / "data"
REPO_BUNDLE_ROOT = ROOT / "assets" / "raw-data" / "各省份"
USER_BUNDLE_ROOT = Path.home() / "Downloads" / "各省份"
DEFAULT_LOCAL_BUNDLE_ROOT = REPO_BUNDLE_ROOT if REPO_BUNDLE_ROOT.exists() else USER_BUNDLE_ROOT
SPECIAL_PLAN_KEYWORDS = [
    "提前",
    "公安",
    "地方专项",
    "高校专项",
    "国家专项",
    "教师专项",
    "卫生专项",
    "免费医学生",
    "定向",
    "少数民族",
    "民族班",
    "预科",
    "综合评价",
    "强基",
    "军校",
    "艺术类",
    "体育类",
]


@dataclass
class AdmissionRecord:
    year: int
    province: str
    track: str
    batch: str
    school_name: str
    school_code: str
    major_group: str
    major_name: str
    plan_type: str
    min_score: int | None
    min_rank: int | None
    admit_count: int | None
    source_url: str
    source_name: str
    notes: str


@dataclass(frozen=True)
class SchoolMetadata:
    school_name: str
    school_code: str
    location: str
    nature: str
    is_985: str
    is_211: str
    city: str


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def as_int(value: str | None) -> int | None:
    if value is None:
        return None
    value = value.strip().replace(",", "")
    if not value:
        return None
    try:
        return int(float(value))
    except ValueError:
        return None


def norm(value: str | None) -> str:
    return (value or "").strip()


def normalize_major_name(value: str | None) -> str:
    text = norm(value)
    text = text.replace("〈", "（").replace("〉", "）").replace("《", "（").replace("》", "）")
    text = re.sub(r"^[/\s]+", "", text)
    text = re.sub(r"^[A-Z]?\d{2,4}[.、\-\s]*", "", text)
    text = re.sub(r"[（(【\[].*$", "", text)
    text = re.sub(r"\d+年?$", "", text)
    text = re.sub(r"\d+$", "", text)
    text = text.replace("（", "").replace("）", "")
    text = re.sub(r"\s+", "", text)
    text = text.strip(" /-—_，,；;:：")
    if text.endswith("师范"):
        text = text[:-2]
    return text


def normalized_header(value: object) -> str:
    value = norm(str(value) if value is not None else "").replace("\u3000", "")
    value = re.sub(r"[（(].*?[）)]", "", value)
    return re.sub(r"\s+", "", value).lower()


def clean_school_name(value: str) -> str:
    value = norm(value)
    value = re.sub(r"[（(](中外合作办学|中外合作|地方专项|国际班|联合培养)[）)]", "", value)
    return value.strip()


def clean_major_group(value: str) -> str:
    value = norm(value)
    return value.strip("（）() ")


def clean_major_name(value: str) -> str:
    value = norm(value)
    value = re.sub(r"^[A-Za-z0-9]{1,4}", "", value).strip()
    value = re.sub(r"[（(].*?[）)]", "", value).strip()
    return value


def infer_plan_type_from_text(text: str) -> str:
    if "中外合作" in text or "中外合办" in text:
        return "中外合作"
    if "联合培养" in text:
        return "联合培养"
    if "地方专项" in text:
        return "地方专项"
    if "国际班" in text or "国际本科" in text:
        return "国际班"
    if "少数民族" in text:
        return "少数民族班"
    if "预科" in text:
        return "预科班"
    return "普通类"


def load_admissions(data_dir: Path) -> list[AdmissionRecord]:
    records: list[AdmissionRecord] = []
    for row in read_csv(data_dir / "admission_records.csv"):
        year = as_int(row.get("year"))
        if year is None:
            continue
        records.append(
            AdmissionRecord(
                year=year,
                province=norm(row.get("province")),
                track=norm(row.get("track")),
                batch=norm(row.get("batch")),
                school_name=norm(row.get("school_name")),
                school_code=norm(row.get("school_code")),
                major_group=norm(row.get("major_group")),
                major_name=norm(row.get("major_name")),
                plan_type=norm(row.get("plan_type")),
                min_score=as_int(row.get("min_score")),
                min_rank=as_int(row.get("min_rank")),
                admit_count=as_int(row.get("admit_count")),
                source_url=norm(row.get("source_url")),
                source_name=norm(row.get("source_name")),
                notes=norm(row.get("notes")),
            )
        )
    return records


def estimate_rank(data_dir: Path, province: str, track: str, score: int | None) -> tuple[int | None, str | None]:
    if score is None:
        return None, None
    rows = [
        row
        for row in read_csv(data_dir / "rank_table.csv")
        if norm(row.get("province")) == province and norm(row.get("track")) == track
    ]
    if not rows:
        return None, None

    parsed = []
    for row in rows:
        row_score = as_int(row.get("score"))
        max_rank = as_int(row.get("max_rank")) or as_int(row.get("min_rank"))
        year = as_int(row.get("year"))
        if row_score is not None and max_rank is not None and year is not None:
            parsed.append((year, row_score, max_rank, norm(row.get("source_name"))))
    if not parsed:
        return None, None

    latest_year = max(item[0] for item in parsed)
    latest = [item for item in parsed if item[0] == latest_year]
    exact = [item for item in latest if item[1] == score]
    if exact:
        item = exact[0]
        return item[2], f"由 {item[0]} 年一分一段表按 {score} 分估算"

    below_or_equal = [item for item in latest if item[1] <= score]
    if below_or_equal:
        item = max(below_or_equal, key=lambda value: value[1])
        return item[2], f"未命中精确分数，按 {item[0]} 年 {item[1]} 分位次保守估算"

    item = min(latest, key=lambda value: value[1])
    return item[2], f"分数低于样本表最低分，按 {item[0]} 年 {item[1]} 分位次估算"


LOCAL_HEADER_ALIASES = {
    "year": ["年份", "年度", "录取年份"],
    "school_name": ["院校名称", "学校名称", "学校", "院校", "招生院校"],
    "school_code": ["院校代码", "学校代码", "学校代号", "院校代号"],
    "track": ["科类", "类别", "首选科目", "考试科类"],
    "batch": ["批次", "录取批次", "批次名称"],
    "major_group": ["所属专业组", "专业组", "专业组代码", "院校专业组", "院校专业组号"],
    "major_name": ["专业", "专业名称", "专业类"],
    "min_score": ["最低分数", "最低分", "投档最低分", "录取最低分", "分数线"],
    "min_rank": ["最低位次", "最低分位", "最低排位", "投档最低排位", "录取最低位次", "位次"],
    "admit_count": ["录取人数", "招生人数", "计划数", "招生计划"],
    "school_location": ["学校所在", "学校省份", "院校所在", "所在省", "省份"],
    "school_nature": ["学校性质", "办学性质", "院校性质", "公私性质"],
    "is_985": ["是否985", "985"],
    "is_211": ["是否211", "211"],
    "city": ["城市", "学校城市", "所在城市"],
    "subject_requirement": ["选科要求", "选科", "科目要求"],
    "major_notes": ["专业备注", "备注", "说明"],
}


def alias_lookup(headers: list[str]) -> dict[str, str]:
    normalized = {normalized_header(header): header for header in headers if header}
    output = {}
    for target, aliases in LOCAL_HEADER_ALIASES.items():
        for alias in aliases:
            match = normalized.get(normalized_header(alias))
            if match:
                output[target] = match
                break
    return output


def find_header_row(matrix: list[list[object]]) -> tuple[int, list[str], dict[str, str]] | None:
    for index, row in enumerate(matrix[:12]):
        headers = [norm(str(cell) if cell is not None else "") for cell in row]
        mapping = alias_lookup(headers)
        has_school = "school_name" in mapping
        has_score = "min_score" in mapping or "min_rank" in mapping
        has_major = "major_name" in mapping or "major_group" in mapping
        if has_school and has_score and has_major:
            return index, headers, mapping
    return None


def year_from_path(path: Path) -> int | None:
    name = path.name
    match = re.search(r"20\d{2}", name)
    if match:
        return int(match.group(0))
    match = re.search(r"(?<!\d)(2[3-9])年", name)
    if match:
        return 2000 + int(match.group(1))
    return None


def local_bundle_candidates(raw_root: Path, province: str, target_years: list[int]) -> list[Path]:
    if not raw_root.exists():
        return []
    year_tokens = {str(year) for year in target_years} | {f"{year % 100}年" for year in target_years}
    candidates = []
    for path in raw_root.rglob("*"):
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        if path.name.startswith(("~$", ".~")):
            continue
        text = path.name
        full_text = str(path)
        if province not in full_text:
            continue
        if not any(token in text for token in year_tokens):
            continue
        if not (("专业" in full_text and "分数" in text) or "录取分数" in text):
            continue
        if any(skip in text for skip in {"招生计划", "一分一段", "省控线", "批次线"}):
            continue
        candidates.append(path)
    return sorted(candidates)


RANK_HEADER_ALIASES = {
    "year": ["年份", "年度"],
    "track": ["科类", "类别", "首选科目"],
    "batch": ["批次", "批次名称"],
    "score": ["分数", "成绩", "总分", "文化总分"],
    "same_score_count": ["本段人数", "同分人数", "人数", "分数段人数"],
    "max_rank": ["累计人数", "累计位次", "本段最低位次", "最低排位", "累计"],
    "rank_range": ["排名区间", "位次区间"],
}


def rank_alias_lookup(headers: list[str]) -> dict[str, str]:
    normalized = {normalized_header(header): header for header in headers if header}
    output = {}
    for target, aliases in RANK_HEADER_ALIASES.items():
        for alias in aliases:
            match = normalized.get(normalized_header(alias))
            if match:
                output[target] = match
                break
    return output


def find_rank_header_row(matrix: list[list[object]]) -> tuple[int, list[str], dict[str, str]] | None:
    for index, row in enumerate(matrix[:12]):
        headers = [norm(str(cell) if cell is not None else "") for cell in row]
        mapping = rank_alias_lookup(headers)
        if "score" in mapping and ("max_rank" in mapping or "rank_range" in mapping):
            return index, headers, mapping
    return None


def local_rank_candidates(raw_root: Path, province: str, target_years: list[int]) -> list[Path]:
    if not raw_root.exists():
        return []
    year_tokens = {str(year) for year in target_years} | {f"{year % 100}年" for year in target_years}
    candidates = []
    for path in raw_root.rglob("*"):
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        if path.name.startswith(("~$", ".~")):
            continue
        text = path.name
        full_text = str(path)
        if province not in full_text:
            continue
        if "一分一段" not in full_text:
            continue
        if not any(token in text for token in year_tokens):
            continue
        candidates.append(path)
    return sorted(candidates)


def parse_score_cell(value: str) -> int | None:
    value = norm(value)
    if not value:
        return None
    match = re.match(r"^(\d+)", value)
    if not match:
        return None
    return int(match.group(1))


def parse_rank_range_max(value: str) -> int | None:
    value = norm(value).replace(",", "").replace("，", "")
    if not value:
        return None
    numbers = [int(item) for item in re.findall(r"\d+", value)]
    if not numbers:
        return None
    return max(numbers)


def load_local_bundle_rank_rows(
    raw_data_root: str | None,
    province: str,
    track: str,
    target_years: list[int],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    if not raw_data_root:
        return [], {"enabled": False, "rows": 0, "files": []}
    raw_root = Path(raw_data_root).expanduser()
    if not raw_root.exists():
        return [], {"enabled": True, "rows": 0, "files": [], "missing_root": str(raw_root)}

    try:
        import openpyxl
    except ModuleNotFoundError:
        return [], {"enabled": True, "rows": 0, "files": [], "error": "missing openpyxl"}

    rows: list[dict[str, object]] = []
    used_files = []
    target_set = set(target_years)
    for path in local_rank_candidates(raw_root, province, target_years):
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        sheet = workbook[workbook.sheetnames[0]]
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
        header_info = find_rank_header_row(matrix)
        if not header_info:
            continue
        header_index, headers, mapping = header_info
        file_year = year_from_path(path)
        file_rows = 0
        for raw in matrix[header_index + 1 :]:
            row = {}
            for index, header in enumerate(headers):
                row[header] = norm(str(raw[index]) if index < len(raw) and raw[index] is not None else "")
            row_track = row.get(mapping.get("track", ""), "") or track
            if norm(row_track) != track:
                continue
            year = as_int(row.get(mapping.get("year", ""), "")) or file_year
            if year is None or (target_set and year not in target_set):
                continue
            score = parse_score_cell(row.get(mapping["score"], ""))
            max_rank = as_int(row.get(mapping.get("max_rank", ""), "")) or parse_rank_range_max(
                row.get(mapping.get("rank_range", ""), "")
            )
            if score is None or max_rank is None:
                continue
            same_score_count = as_int(row.get(mapping.get("same_score_count", ""), ""))
            rows.append(
                {
                    "year": year,
                    "province": province,
                    "track": track,
                    "score": score,
                    "max_rank": max_rank,
                    "same_score_count": same_score_count,
                    "source_name": "本地一分一段Excel",
                    "source_path": str(path),
                }
            )
            file_rows += 1
        if file_rows:
            used_files.append({"path": str(path), "rows": file_rows})
    return rows, {
        "enabled": True,
        "rows": len(rows),
        "files": used_files,
        "root": str(raw_root),
        "years": sorted({int(row["year"]) for row in rows}),
    }


def estimate_rank_from_local_rows(
    local_rank_rows: list[dict[str, object]],
    score: int | None,
) -> tuple[int | None, str | None]:
    if score is None or not local_rank_rows:
        return None, None
    parsed = [
        (int(row["year"]), int(row["score"]), int(row["max_rank"]), str(row.get("source_name") or "本地一分一段Excel"))
        for row in local_rank_rows
        if row.get("year") is not None and row.get("score") is not None and row.get("max_rank") is not None
    ]
    if not parsed:
        return None, None
    latest_year = max(item[0] for item in parsed)
    latest = [item for item in parsed if item[0] == latest_year]
    exact = [item for item in latest if item[1] == score]
    if exact:
        item = exact[0]
        return item[2], f"由本地 {item[0]} 年一分一段表按 {score} 分估算"
    below_or_equal = [item for item in latest if item[1] <= score]
    if below_or_equal:
        item = max(below_or_equal, key=lambda value: value[1])
        return item[2], f"未命中精确分数，按本地 {item[0]} 年 {item[1]} 分位次保守估算"
    item = min(latest, key=lambda value: value[1])
    return item[2], f"分数低于本地样本表最低分，按 {item[0]} 年 {item[1]} 分位次估算"


def build_metadata_key(school_code: str, school_name: str) -> tuple[str, str]:
    return norm(school_code), clean_school_name(school_name)


def add_metadata(
    metadata: dict[tuple[str, str], SchoolMetadata],
    school_name: str,
    school_code: str,
    location: str,
    nature: str,
    is_985: str,
    is_211: str,
    city: str,
) -> None:
    school_name = clean_school_name(school_name)
    school_code = norm(school_code)
    if not school_name:
        return
    item = SchoolMetadata(
        school_name=school_name,
        school_code=school_code,
        location=norm(location),
        nature=norm(nature),
        is_985=norm(is_985),
        is_211=norm(is_211),
        city=norm(city),
    )
    metadata[(school_code, school_name)] = item
    metadata[("", school_name)] = item


def metadata_for(record: dict[str, object], metadata: dict[tuple[str, str], SchoolMetadata]) -> SchoolMetadata | None:
    school_code = norm(str(record.get("school_code") or ""))
    school_name = clean_school_name(str(record.get("school_name") or ""))
    return metadata.get((school_code, school_name)) or metadata.get(("", school_name))


def load_local_bundle_admissions(
    raw_data_root: str | None,
    province: str,
    track: str,
    target_years: list[int],
) -> tuple[list[AdmissionRecord], dict[tuple[str, str], SchoolMetadata], dict[str, object]]:
    if not raw_data_root:
        return [], {}, {"enabled": False, "rows": 0, "files": []}
    raw_root = Path(raw_data_root).expanduser()
    if not raw_root.exists():
        return [], {}, {"enabled": True, "rows": 0, "files": [], "missing_root": str(raw_root)}

    try:
        import openpyxl
    except ModuleNotFoundError:
        return [], {}, {"enabled": True, "rows": 0, "files": [], "error": "missing openpyxl"}

    records: list[AdmissionRecord] = []
    metadata: dict[tuple[str, str], SchoolMetadata] = {}
    used_files = []
    target_set = set(target_years)
    for path in local_bundle_candidates(raw_root, province, target_years):
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        sheet = workbook[workbook.sheetnames[0]]
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
        header_info = find_header_row(matrix)
        if not header_info:
            continue
        header_index, headers, mapping = header_info
        file_year = year_from_path(path)
        file_rows = 0
        for raw in matrix[header_index + 1 :]:
            row = {}
            for index, header in enumerate(headers):
                row[header] = norm(str(raw[index]) if index < len(raw) and raw[index] is not None else "")
            row_track = row.get(mapping.get("track", ""), "") or track
            if norm(row_track) != track:
                continue
            year = as_int(row.get(mapping.get("year", ""), "")) or file_year
            if year is None or (target_set and year not in target_set):
                continue
            school_name = row.get(mapping.get("school_name", ""), "")
            min_score = as_int(row.get(mapping.get("min_score", ""), ""))
            min_rank = as_int(row.get(mapping.get("min_rank", ""), ""))
            if not school_name or (min_score is None and min_rank is None):
                continue
            school_code = row.get(mapping.get("school_code", ""), "")
            major_group = clean_major_group(row.get(mapping.get("major_group", ""), ""))
            major_name = clean_major_name(row.get(mapping.get("major_name", ""), ""))
            batch = row.get(mapping.get("batch", ""), "") or "本科批"
            admit_count = as_int(row.get(mapping.get("admit_count", ""), ""))
            subject_requirement = row.get(mapping.get("subject_requirement", ""), "")
            major_notes = row.get(mapping.get("major_notes", ""), "")
            location = row.get(mapping.get("school_location", ""), "")
            nature = row.get(mapping.get("school_nature", ""), "")
            is_985 = row.get(mapping.get("is_985", ""), "")
            is_211 = row.get(mapping.get("is_211", ""), "")
            city = row.get(mapping.get("city", ""), "")
            add_metadata(metadata, school_name, school_code, location, nature, is_985, is_211, city)
            row_text = " ".join(value for value in row.values() if value)
            plan_type = infer_plan_type_from_text(row_text)
            notes = "；".join(
                part
                for part in [
                    f"选科：{subject_requirement}" if subject_requirement else "",
                    major_notes,
                    f"学校所在：{location}" if location else "",
                    f"学校性质：{nature}" if nature else "",
                    f"985：{is_985}" if is_985 else "",
                    f"211：{is_211}" if is_211 else "",
                    f"本地表格：{path.name}",
                ]
                if part
            )
            records.append(
                AdmissionRecord(
                    year=year,
                    province=province,
                    track=track,
                    batch=batch,
                    school_name=clean_school_name(school_name),
                    school_code=norm(school_code),
                    major_group=major_group,
                    major_name=major_name,
                    plan_type=plan_type,
                    min_score=min_score,
                    min_rank=min_rank,
                    admit_count=admit_count,
                    source_url=str(path),
                    source_name="本地专业分数线Excel",
                    notes=notes,
                )
            )
            file_rows += 1
        if file_rows:
            used_files.append({"path": str(path), "rows": file_rows})

    return records, metadata, {"enabled": True, "rows": len(records), "files": used_files, "root": str(raw_root)}


def median_int(values: Iterable[int]) -> int | None:
    values = list(values)
    if not values:
        return None
    return int(round(statistics.median(values)))


def dynamic_rank_band(rank: int) -> int:
    return max(1000, min(15000, int(rank * 0.08)))


def option_key(record: AdmissionRecord) -> tuple[str, str, str, str, str, str]:
    return (
        record.school_name,
        record.school_code,
        record.batch,
        record.major_group,
        record.major_name,
        record.plan_type,
    )


def data_precision(major_group: str, major_name: str) -> str:
    if major_name:
        return "专业级"
    if major_group:
        return "专业组级"
    return "学校级"


def is_special_plan(record: AdmissionRecord) -> bool:
    text = " ".join(
        [
            record.batch,
            record.plan_type,
            record.major_group,
            record.major_name,
            record.school_name if "专项" in record.school_name else "",
        ]
    )
    return any(keyword in text for keyword in SPECIAL_PLAN_KEYWORDS)


def classify_by_rank(candidate_rank: int, cutoff_rank: int) -> tuple[str, int, str]:
    band = dynamic_rank_band(candidate_rank)
    gap = cutoff_rank - candidate_rank
    if gap >= band:
        return "保", gap, f"位次优于历史中位线约 {gap} 名"
    if gap >= 0:
        return "稳", gap, f"位次略优于历史中位线约 {gap} 名"
    if gap >= -band:
        return "冲", gap, f"位次略低于历史中位线约 {abs(gap)} 名"
    return "险", gap, f"位次低于历史中位线约 {abs(gap)} 名"


def classify_by_score(candidate_score: int, cutoff_score: int) -> tuple[str, int, str]:
    gap = candidate_score - cutoff_score
    if gap >= 10:
        return "保", gap, f"分数高于历史中位线约 {gap} 分"
    if gap >= 0:
        return "稳", gap, f"分数略高于历史中位线约 {gap} 分"
    if gap >= -10:
        return "冲", gap, f"分数略低于历史中位线约 {abs(gap)} 分"
    return "险", gap, f"分数低于历史中位线约 {abs(gap)} 分"


def build_recommendations(
    records: list[AdmissionRecord],
    province: str,
    track: str,
    score: int | None,
    rank: int | None,
    include_risky: bool,
    include_special_plans: bool,
) -> list[dict[str, object]]:
    filtered = [
        item
        for item in records
        if item.province == province and item.track == track and item.school_name
    ]
    if not include_special_plans:
        filtered = [item for item in filtered if not is_special_plan(item)]
    prefer_rank = rank is not None and any(item.min_rank is not None for item in filtered)
    grouped: dict[tuple[str, str, str, str, str, str], list[AdmissionRecord]] = {}
    for item in filtered:
        grouped.setdefault(option_key(item), []).append(item)

    recommendations: list[dict[str, object]] = []
    for key, items in grouped.items():
        ranks = [item.min_rank for item in items if item.min_rank is not None]
        scores = [item.min_score for item in items if item.min_score is not None]
        cutoff_rank = median_int(ranks)
        cutoff_score = median_int(scores)

        if prefer_rank and cutoff_rank is None:
            continue
        if rank is not None and cutoff_rank is not None:
            level, gap, note = classify_by_rank(rank, cutoff_rank)
            basis = "rank"
        elif score is not None and cutoff_score is not None:
            level, gap, note = classify_by_score(score, cutoff_score)
            basis = "score"
        else:
            continue

        if level == "险" and not include_risky:
            continue

        years = sorted({item.year for item in items})
        source_names = sorted({item.source_name for item in items if item.source_name})
        demo = any("DEMO" in item.source_name.upper() or item.school_name.startswith("示例") for item in items)
        recommendations.append(
            {
                "level": level,
                "school_name": key[0],
                "school_code": key[1],
                "batch": key[2],
                "major_group": key[3],
                "major_name": key[4],
                "plan_type": key[5],
                "median_rank": cutoff_rank,
                "median_score": cutoff_score,
                "gap": gap,
                "basis": basis,
                "years": years,
                "note": note,
                "source_names": source_names,
                "demo": demo,
                "record_count": len(items),
                "data_precision": data_precision(key[3], key[4]),
            }
        )

    level_order = {"冲": 0, "稳": 1, "保": 2, "险": 3}

    def sort_key(item: dict[str, object]) -> tuple[int, int]:
        gap = int(item["gap"])
        level = str(item["level"])
        if level == "冲":
            closeness = abs(gap)
        elif level in {"稳", "保"}:
            closeness = gap
        else:
            closeness = abs(gap)
        return level_order.get(level, 9), closeness

    return sorted(recommendations, key=sort_key)


def split_interests(raw: str) -> list[str]:
    parts: list[str] = []
    for chunk in raw.replace("，", ",").replace("、", ",").replace("/", ",").split(","):
        chunk = chunk.strip()
        if chunk:
            parts.append(chunk.lower())
    return parts


INTEREST_SYNONYMS = {
    "ai": ["人工智能", "智能科学", "机器学习", "ai", "计算机", "软件", "数据", "自动化"],
    "人工智能": ["人工智能", "智能科学", "机器学习", "ai", "计算机", "软件", "数据", "自动化"],
    "计算机": ["计算机", "软件", "人工智能", "数据", "大数据", "网络安全", "信息安全", "物联网"],
    "财经": ["财经", "金融", "经济", "会计", "财务", "财政", "投资"],
    "医学": ["医学", "临床", "口腔", "药学", "护理", "生物医学"],
    "新能源": ["新能源", "能源", "电气", "储能", "材料", "车辆"],
    "传媒": ["传媒", "新闻", "传播", "广告", "网络与新媒体"],
}


def expanded_interest_terms(interests_raw: str) -> list[str]:
    terms: list[str] = []
    for interest in split_interests(interests_raw):
        terms.append(interest)
        terms.extend(INTEREST_SYNONYMS.get(interest, []))
    seen = set()
    output = []
    for term in terms:
        key = term.lower()
        if key not in seen:
            seen.add(key)
            output.append(key)
    return output


def major_matches_interest(major_name: str, interests_raw: str) -> bool:
    major = major_name.lower()
    return bool(major and any(term and term in major for term in expanded_interest_terms(interests_raw)))


SPECIAL_CONTROLLED_TERMS = ["警官", "公安", "司法", "刑事", "海关", "军校", "军队", "执法", "治安", "侦查"]
SPECIAL_INTEREST_TERMS = ["公安", "警察", "警校", "司法", "海关", "军校", "政法"]


def user_allows_special_controlled_options(interests_raw: str) -> bool:
    return any(term in interests_raw for term in SPECIAL_INTEREST_TERMS)


def is_special_controlled_option(item: dict[str, object]) -> bool:
    text = " ".join(
        str(item.get(key) or "")
        for key in ["school_name", "major_name", "major_group", "plan_type", "notes"]
    )
    return any(term in text for term in SPECIAL_CONTROLLED_TERMS)


def match_majors(data_dir: Path, interests_raw: str, limit: int) -> list[dict[str, str]]:
    interests = split_interests(interests_raw)
    if not interests:
        return []

    matches: list[tuple[int, dict[str, str]]] = []
    for row in read_csv(data_dir / "majors.csv"):
        major_name = norm(row.get("major_name"))
        major_key = major_name.lower()
        keywords = [item.strip().lower() for item in norm(row.get("interest_keywords")).split("|") if item.strip()]
        score = 0
        for interest in interests:
            if interest and (interest in major_key or major_key == interest):
                score += 10
            elif interest in keywords:
                score += 3
            elif any(interest in keyword or keyword in interest for keyword in keywords):
                score += 1
        if score > 0:
            matches.append((score, {key: norm(value) for key, value in row.items()}))

    matches.sort(key=lambda item: item[0], reverse=True)
    return [item[1] for item in matches[:limit]]


def load_major_profiles(data_dir: Path) -> dict[str, dict[str, str]]:
    profiles: dict[str, dict[str, str]] = {}
    for row in read_csv(data_dir / "majors.csv"):
        key = normalize_major_name(row.get("major_name"))
        if key:
            profiles[key] = {field: norm(value) for field, value in row.items()}
    return profiles


def major_profile_note(major_name: str, profiles: dict[str, dict[str, str]]) -> str:
    key = normalize_major_name(major_name)
    if not key:
        return ""
    profile = profiles.get(key)
    if not profile:
        for candidate_key, candidate in profiles.items():
            if key in candidate_key or candidate_key in key:
                profile = candidate
                break
    if not profile:
        return ""
    outlook = profile.get("employment_outlook", "")
    roles = profile.get("typical_roles", "")
    if roles:
        return f"{outlook}；常见去向：{roles}"
    return outlook


def attach_major_profiles(
    recommendations: list[dict[str, object]],
    profiles: dict[str, dict[str, str]],
) -> list[dict[str, object]]:
    for item in recommendations:
        major_name = str(item.get("major_name") or "")
        item["major_profile"] = major_profile_note(major_name, profiles) if major_name else ""
    return recommendations


def parse_target_years(raw: str) -> list[int]:
    years: list[int] = []
    for chunk in raw.replace("，", ",").split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        try:
            years.append(int(chunk))
        except ValueError as exc:
            raise SystemExit(f"目标年份格式错误：{chunk}") from exc
    return sorted(set(years))


def collect_rank_years(data_dir: Path, province: str, track: str) -> tuple[list[int], int]:
    rows = [
        row
        for row in read_csv(data_dir / "rank_table.csv")
        if norm(row.get("province")) == province and norm(row.get("track")) == track
    ]
    years = sorted({year for row in rows if (year := as_int(row.get("year"))) is not None})
    return years, len(rows)


def collect_manifest_gaps(data_dir: Path, province: str, track: str, target_years: list[int]) -> list[dict[str, str]]:
    target_set = set(target_years)
    gaps: list[dict[str, str]] = []
    for row in read_csv(data_dir / "collection_manifest.csv"):
        row_province = norm(row.get("province"))
        row_track = norm(row.get("track"))
        if row_province != province or (row_track and row_track != track):
            continue
        year = as_int(row.get("year"))
        if target_set and year not in target_set:
            continue
        status = norm(row.get("status"))
        if status in {"needs_discovery", "needs_ocr"}:
            gaps.append(
                {
                    "year": str(year or ""),
                    "dataset": norm(row.get("dataset")),
                    "status": status,
                    "title": norm(row.get("article_title")),
                }
            )
    return gaps


def collect_coverage(
    records: list[AdmissionRecord],
    data_dir: Path,
    province: str,
    track: str,
    target_years: list[int],
    local_rank_years: list[int] | None = None,
    local_rank_rows: int = 0,
) -> dict[str, object]:
    selected = [item for item in records if item.province == province and item.track == track]
    admission_years = sorted({item.year for item in selected})
    rank_years, rank_rows = collect_rank_years(data_dir, province, track)
    rank_years = sorted(set(rank_years) | set(local_rank_years or []))
    rank_rows += local_rank_rows
    target_set = set(target_years)
    major_level_rows = sum(1 for item in selected if item.major_name)
    major_group_only_rows = sum(1 for item in selected if item.major_group and not item.major_name)
    school_only_rows = sum(1 for item in selected if not item.major_group and not item.major_name)
    return {
        "province": province,
        "track": track,
        "target_years": target_years,
        "years": admission_years,
        "rank_years": rank_years,
        "rows": len(selected),
        "rank_rows": rank_rows,
        "major_level_rows": major_level_rows,
        "major_group_only_rows": major_group_only_rows,
        "school_only_rows": school_only_rows,
        "missing_admission_years": sorted(target_set - set(admission_years)) if target_set else [],
        "missing_rank_years": sorted(target_set - set(rank_years)) if target_set else [],
        "manifest_gaps": collect_manifest_gaps(data_dir, province, track, target_years),
        "demo_rows": sum(
            1
            for item in selected
            if "DEMO" in item.source_name.upper() or item.school_name.startswith("示例")
        ),
    }


def grouped_by_level(
    recommendations: list[dict[str, object]],
    per_level: int,
    interests_raw: str = "",
) -> dict[str, list[dict[str, object]]]:
    grouped: dict[str, list[dict[str, object]]] = {"冲": [], "稳": [], "保": [], "险": []}
    if interests_raw:
        recommendations = sorted(recommendations, key=lambda item: recommendation_priority(item, interests_raw))
    for item in recommendations:
        level = str(item["level"])
        if len(grouped.setdefault(level, [])) < per_level:
            grouped[level].append(item)
    return grouped


def filter_interest_admission_matches(
    recommendations: list[dict[str, object]],
    interests_raw: str,
    limit: int,
) -> list[dict[str, object]]:
    if not interests_raw:
        return []
    allow_special = user_allows_special_controlled_options(interests_raw)
    matches = [
        item
        for item in recommendations
        if str(item.get("major_name") or "") and major_matches_interest(str(item.get("major_name") or ""), interests_raw)
        and (allow_special or not is_special_controlled_option(item))
    ]
    level_order = {"冲": 0, "稳": 1, "保": 2, "险": 3}
    matches.sort(key=lambda item: (level_order.get(str(item.get("level")), 9), abs(int(item.get("gap") or 0))))
    return matches[:limit]


def attach_school_metadata(
    recommendations: list[dict[str, object]],
    metadata: dict[tuple[str, str], SchoolMetadata],
) -> None:
    for item in recommendations:
        meta = metadata_for(item, metadata)
        if not meta:
            continue
        item["school_location"] = meta.location
        item["school_nature"] = meta.nature
        item["is_985"] = meta.is_985
        item["is_211"] = meta.is_211
        item["school_city"] = meta.city


def school_info_text(item: dict[str, object]) -> str:
    parts = []
    location = norm(str(item.get("school_location") or ""))
    city = norm(str(item.get("school_city") or ""))
    nature = norm(str(item.get("school_nature") or ""))
    is_985 = norm(str(item.get("is_985") or ""))
    is_211 = norm(str(item.get("is_211") or ""))
    if location:
        parts.append(location + (f"/{city}" if city and city != location else ""))
    if nature:
        parts.append(nature)
    if is_985 == "是":
        parts.append("985")
    if is_211 == "是":
        parts.append("211")
    return " / ".join(parts) or "-"


def option_detail_text(item: dict[str, object]) -> str:
    detail_parts = []
    if item.get("batch"):
        detail_parts.append(str(item["batch"]))
    if item.get("major_group"):
        detail_parts.append(f"专业组 {item['major_group']}")
    if item.get("major_name"):
        detail_parts.append(str(item["major_name"]))
    if item.get("plan_type"):
        detail_parts.append(str(item["plan_type"]))
    return " / ".join(detail_parts) or "-"


def recommendation_priority(item: dict[str, object], interests_raw: str) -> tuple[int, int, int, int]:
    level_priority = {"稳": 0, "保": 1, "冲": 2, "险": 3}
    level = str(item.get("level") or "")
    gap = int(item.get("gap") or 0)
    major_name = str(item.get("major_name") or "")
    interest_penalty = 0 if major_matches_interest(major_name, interests_raw) else 1
    plan_penalty = 1 if str(item.get("plan_type") or "") in {"地方专项", "中外合作", "国际班"} else 0
    controlled_penalty = 20 if (
        is_special_controlled_option(item) and not user_allows_special_controlled_options(interests_raw)
    ) else 0
    if level == "冲":
        distance = abs(gap)
    elif level in {"稳", "保"}:
        distance = gap
    else:
        distance = abs(gap)
    return (controlled_penalty + interest_penalty, level_priority.get(level, 9), plan_penalty, distance)


def pick_diverse_recommendations(
    items: list[dict[str, object]],
    interests_raw: str,
    limit: int,
) -> list[dict[str, object]]:
    selected = []
    seen_schools = set()
    for item in sorted(items, key=lambda value: recommendation_priority(value, interests_raw)):
        school = str(item.get("school_name") or "")
        if school in seen_schools:
            continue
        selected.append(item)
        seen_schools.add(school)
        if len(selected) >= limit:
            break
    return selected


def build_regional_recommendations(
    recommendations: list[dict[str, object]],
    home_province: str,
    interests_raw: str,
    home_limit: int,
    away_limit: int,
) -> dict[str, list[dict[str, object]]]:
    if not home_province:
        return {"home": [], "away": []}
    with_location = [item for item in recommendations if norm(str(item.get("school_location") or ""))]
    home = [item for item in with_location if norm(str(item.get("school_location") or "")) == home_province]
    away = [item for item in with_location if norm(str(item.get("school_location") or "")) != home_province]
    return {
        "home": pick_diverse_recommendations(home, interests_raw, home_limit),
        "away": pick_diverse_recommendations(away, interests_raw, away_limit),
    }


def markdown_table(rows: list[list[str]], headers: list[str]) -> str:
    if not rows:
        return ""
    output = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        output.append("| " + " | ".join(cell.replace("\n", " ") for cell in row) + " |")
    return "\n".join(output)


def render_markdown(result: dict[str, object]) -> str:
    lines: list[str] = ["# 高考志愿参考建议", ""]
    input_summary = result["input"]
    coverage = result["coverage"]
    lines.append("## 输入")
    lines.append(
        f"- 省份：{input_summary['province']}；科类/选科：{input_summary['track']}；"
        f"分数：{input_summary.get('score') or '未提供'}；位次：{input_summary.get('rank') or '未提供'}"
    )
    if input_summary.get("rank_note"):
        lines.append(f"- 位次说明：{input_summary['rank_note']}")
    if input_summary.get("interests"):
        lines.append(f"- 兴趣方向：{input_summary['interests']}")
    if input_summary.get("raw_data_root"):
        lines.append(f"- 本地原始数据目录：{input_summary['raw_data_root']}")
    lines.append("")

    lines.append("## 数据覆盖")
    years = coverage.get("years") or []
    rank_years = coverage.get("rank_years") or []
    target_years = coverage.get("target_years") or []
    if target_years:
        lines.append(f"- 历史目标年份：{', '.join(str(year) for year in target_years)}")
    lines.append(
        f"- 录取线数据：{coverage['province']} / {coverage['track']}，年份："
        f"{', '.join(str(year) for year in years) if years else '无'}，记录数：{coverage['rows']}"
    )
    lines.append(
        f"- 一分一段数据：年份："
        f"{', '.join(str(year) for year in rank_years) if rank_years else '无'}，记录数：{coverage.get('rank_rows', 0)}"
    )
    lines.append(
        f"- 录取线细度：专业级 {coverage.get('major_level_rows', 0)} 条；"
        f"院校专业组级 {coverage.get('major_group_only_rows', 0)} 条；学校级 {coverage.get('school_only_rows', 0)} 条"
    )
    if input_summary.get("include_special_plans"):
        lines.append("- 推荐范围：已包含提前批、专项、定向、艺术体育等特殊计划；请逐项核对资格限制。")
    else:
        lines.append("- 推荐范围：默认排除提前批、专项、定向、艺术体育等特殊计划；需要研究特殊批次时使用 `--include-special-plans`。")
    if coverage.get("missing_admission_years"):
        missing = ", ".join(str(year) for year in coverage["missing_admission_years"])
        lines.append(f"- 缺录取线年份：{missing}。当前结果只能做阶段性参考。")
    if coverage.get("missing_rank_years"):
        missing = ", ".join(str(year) for year in coverage["missing_rank_years"])
        lines.append(f"- 缺一分一段年份：{missing}。位次换算和跨年比较仍需补齐。")
    if coverage.get("manifest_gaps"):
        gap_text = []
        for item in coverage["manifest_gaps"]:
            gap_text.append(f"{item['year']} {item['dataset']}({item['status']})")
        lines.append(f"- 采集清单未完成：{'; '.join(gap_text)}。")
    if coverage.get("demo_rows"):
        lines.append(f"- 注意：当前命中 {coverage['demo_rows']} 条样例数据，不能用于真实填报。")
    local_bundle = coverage.get("local_bundle") or {}
    if local_bundle.get("enabled"):
        if local_bundle.get("rows"):
            files = local_bundle.get("files") or []
            lines.append(f"- 本地 Excel 专业线：已合并 {local_bundle['rows']} 条，来自 {len(files)} 个文件。")
        elif local_bundle.get("missing_root"):
            lines.append(f"- 本地 Excel 专业线：目录不存在，未合并：{local_bundle['missing_root']}")
        elif local_bundle.get("error"):
            lines.append(f"- 本地 Excel 专业线：未合并，原因：{local_bundle['error']}")
        else:
            lines.append("- 本地 Excel 专业线：没有识别到可合并的目标年份专业录取表。")
    local_rank_bundle = coverage.get("local_rank_bundle") or {}
    if local_rank_bundle.get("enabled"):
        if local_rank_bundle.get("rows"):
            files = local_rank_bundle.get("files") or []
            years_text = ", ".join(str(year) for year in local_rank_bundle.get("years", []))
            lines.append(f"- 本地一分一段表：已读取 {local_rank_bundle['rows']} 行，年份：{years_text or '未知'}，来自 {len(files)} 个文件。")
        elif local_rank_bundle.get("missing_root"):
            lines.append(f"- 本地一分一段表：目录不存在，未读取：{local_rank_bundle['missing_root']}")
        elif local_rank_bundle.get("error"):
            lines.append(f"- 本地一分一段表：未读取，原因：{local_rank_bundle['error']}")
        else:
            lines.append("- 本地一分一段表：没有识别到可用表格；只给分数时会退回分数线粗判。")
    if not coverage.get("rows"):
        lines.append("- 当前数据目录没有命中记录，请先导入该省份、科类、近三年录取数据。")

    all_recommendations = [
        item
        for level_items in result["recommendations_by_level"].values()
        for item in level_items
    ]
    shown_admission_matches = result.get("admission_major_matches") or []
    shown_records = [*all_recommendations, *shown_admission_matches]
    if any("掌上高考" in ",".join(item.get("source_names", [])) for item in shown_records):
        lines.append("- 注意：部分命中结果来自掌上高考聚合补充源，已标注来源，正式填报前必须用学校官网或招生计划复核。")
    if any("提前" in str(item.get("batch") or "") or "公安" in str(item.get("plan_type") or "") for item in shown_records):
        lines.append("- 注意：当前结果含提前批/公安等特殊批次，政审、体检、体测、性别、地市和选科限制不能按普通本科批直接比较。")
    lines.append("")

    regional = result.get("regional_recommendations") or {}
    regional_rows = []
    for label, items in [("本省重点", regional.get("home") or []), ("省外重点", regional.get("away") or [])]:
        for item in items:
            regional_rows.append(
                [
                    label,
                    str(item.get("school_name") or "-"),
                    option_detail_text(item),
                    str(item.get("level") or "-"),
                    str(item.get("median_rank") or "-"),
                    str(item.get("median_score") or "-"),
                    school_info_text(item),
                    str(item.get("note") or "-"),
                ]
            )
    if regional_rows:
        lines.append("## 本省/省外重点")
        lines.append(
            markdown_table(
                regional_rows,
                ["范围", "院校", "专业/专业组", "档位", "历史中位位次", "历史中位分", "学校信息", "判断"],
            )
        )
        lines.append("")

    grouped = result["recommendations_by_level"]
    for level in ["冲", "稳", "保", "险"]:
        items = grouped.get(level) or []
        if not items:
            continue
        rows = []
        for item in items:
            rows.append(
                [
                    str(item["school_name"]),
                    option_detail_text(item),
                    str(item.get("median_rank") or "-"),
                    str(item.get("median_score") or "-"),
                    str(item.get("note") or "-"),
                    str(item.get("data_precision") or "-"),
                    ", ".join(str(year) for year in item.get("years", [])),
                    ", ".join(item.get("source_names", [])) or "-",
                ]
            )
        lines.append(f"## {level} 档")
        lines.append(markdown_table(rows, ["院校", "专业组/专业", "历史中位位次", "历史中位分", "判断", "数据精度", "年份", "来源"]))
        lines.append("")

    major_matches = result["major_matches"]
    admission_major_matches = result.get("admission_major_matches") or []
    if admission_major_matches:
        rows = []
        for item in admission_major_matches:
            type_parts = []
            if item.get("batch"):
                type_parts.append(str(item["batch"]))
            if item.get("major_group"):
                type_parts.append(f"专业组 {item['major_group']}")
            if item.get("plan_type"):
                type_parts.append(str(item["plan_type"]))
            rows.append(
                [
                    str(item.get("school_name") or "-"),
                    str(item.get("major_name") or "-"),
                    " / ".join(type_parts) or "-",
                    str(item.get("level") or "-"),
                    str(item.get("median_rank") or "-"),
                    str(item.get("median_score") or "-"),
                    str(item.get("note") or "-"),
                    str(item.get("data_precision") or "-"),
                    str(item.get("major_profile") or "-"),
                    ", ".join(str(year) for year in item.get("years", [])),
                    ", ".join(item.get("source_names", [])) or "-",
                ]
            )
        lines.append("## 已导入专业录取线匹配")
        lines.append(markdown_table(rows, ["院校", "专业", "批次/类型/专业组", "档位", "历史中位位次", "历史中位分", "判断", "数据精度", "专业参考", "年份", "来源"]))
        lines.append("")

    if major_matches:
        rows = []
        for item in major_matches:
            rows.append(
                [
                    item.get("major_name", "-"),
                    item.get("degree_category", "-"),
                    item.get("employment_outlook", "-"),
                    item.get("typical_roles", "-"),
                    item.get("risk_notes", "-"),
                ]
            )
        lines.append("## 专业方向")
        lines.append(markdown_table(rows, ["专业", "门类", "就业前景", "常见去向", "风险点"]))
        lines.append("")
    elif input_summary.get("interests"):
        lines.append("## 专业方向")
        lines.append("- 当前 majors.csv 没有匹配到兴趣关键词，请补充专业库或换更具体的方向。")
        lines.append("")

    lines.append("## 下一步必须核对")
    lines.append("- 2026 年本省一分一段表、批次线、志愿模式和投档规则。")
    lines.append("- 目标院校 2026 招生计划、专业组、选科要求、学费、校区和中外合作/专项计划差异。")
    lines.append("- 考生城市偏好、家庭预算、是否接受民办/独立学院/中外合作、读研或就业倾向。")
    return "\n".join(lines)


def build_result(args: argparse.Namespace) -> dict[str, object]:
    data_dir = Path(args.data_dir).expanduser().resolve()
    target_years = parse_target_years(args.target_years)
    admissions = load_admissions(data_dir)
    include_special_plans = getattr(args, "include_special_plans", False)
    raw_data_root = getattr(args, "raw_data_root", "") or ""
    local_admissions, school_metadata, local_bundle = load_local_bundle_admissions(
        raw_data_root,
        args.province,
        args.track,
        target_years,
    )
    local_rank_rows, local_rank_bundle = load_local_bundle_rank_rows(
        raw_data_root,
        args.province,
        args.track,
        target_years,
    )
    admissions.extend(local_admissions)
    rank = args.rank
    rank_note = None
    if rank is None:
        rank, rank_note = estimate_rank(data_dir, args.province, args.track, args.score)
    if rank is None:
        rank, rank_note = estimate_rank_from_local_rows(local_rank_rows, args.score)

    recommendations = build_recommendations(
        admissions,
        args.province,
        args.track,
        args.score,
        rank,
        args.include_risky,
        include_special_plans,
    )
    major_profiles = load_major_profiles(data_dir)
    recommendations = attach_major_profiles(recommendations, major_profiles)
    attach_school_metadata(recommendations, school_metadata)
    home_province = getattr(args, "home_province", "") or args.province
    result = {
        "input": {
            "province": args.province,
            "track": args.track,
            "score": args.score,
            "rank": rank,
            "rank_note": rank_note,
            "interests": args.interests,
            "data_dir": str(data_dir),
            "include_special_plans": include_special_plans,
            "raw_data_root": raw_data_root,
            "home_province": home_province,
        },
        "coverage": collect_coverage(
            admissions,
            data_dir,
            args.province,
            args.track,
            target_years,
            local_rank_years=[int(year) for year in local_rank_bundle.get("years", [])],
            local_rank_rows=int(local_rank_bundle.get("rows", 0) or 0),
        ),
        "recommendations_by_level": grouped_by_level(recommendations, args.per_level, args.interests or ""),
        "regional_recommendations": build_regional_recommendations(
            recommendations,
            home_province,
            args.interests or "",
            getattr(args, "home_limit", 1),
            getattr(args, "away_limit", 5),
        ),
        "admission_major_matches": filter_interest_admission_matches(
            recommendations,
            args.interests or "",
            args.major_limit,
        ),
        "major_matches": match_majors(data_dir, args.interests or "", args.major_limit),
    }
    result["coverage"]["local_bundle"] = local_bundle
    result["coverage"]["local_rank_bundle"] = local_rank_bundle
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Recommend Gaokao colleges and majors from CSV data.")
    parser.add_argument("--province", required=True, help="Candidate province, e.g. 广东")
    parser.add_argument("--track", required=True, help="Subject type, e.g. 物理类")
    parser.add_argument("--score", type=int, help="Gaokao score")
    parser.add_argument("--rank", type=int, help="Candidate rank/位次")
    parser.add_argument("--interests", default="", help="Comma-separated interests, e.g. 人工智能,新能源")
    parser.add_argument("--data-dir", default=str(DEFAULT_DATA_DIR), help="Directory containing normalized CSV files")
    parser.add_argument(
        "--raw-data-root",
        default=str(DEFAULT_LOCAL_BUNDLE_ROOT) if DEFAULT_LOCAL_BUNDLE_ROOT.exists() else "",
        help="Optional raw Excel bundle root, defaults to assets/raw-data/各省份 when bundled, otherwise ~/Downloads/各省份",
    )
    parser.add_argument("--home-province", help="Candidate's home province for in-province/out-of-province picks")
    parser.add_argument("--home-limit", type=int, default=1, help="Number of in-province highlight picks")
    parser.add_argument("--away-limit", type=int, default=5, help="Number of out-of-province highlight picks")
    parser.add_argument("--per-level", type=int, default=8, help="Maximum recommendations per level")
    parser.add_argument("--major-limit", type=int, default=8, help="Maximum major suggestions")
    parser.add_argument("--include-risky", action="store_true", help="Include high-risk options")
    parser.add_argument("--include-special-plans", action="store_true", help="Include提前批、专项、定向、艺术体育等特殊计划")
    parser.add_argument("--target-years", default="2023,2024,2025", help="Historical baseline years")
    parser.add_argument("--format", choices=["markdown", "json"], default="markdown")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.score is None and args.rank is None:
        raise SystemExit("请至少提供 --score 或 --rank。")
    result = build_result(args)
    if args.format == "json":
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(render_markdown(result))


if __name__ == "__main__":
    main()
