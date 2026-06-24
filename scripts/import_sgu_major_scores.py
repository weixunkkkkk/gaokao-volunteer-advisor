#!/usr/bin/env python3
"""Import Shaoguan University official 2025 Guangdong major-level scores."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
SOURCE_NAME = "韶关学院招生信息网"
SCHOOL_NAME = "韶关学院"
SCHOOL_CODE = "10576"
SOURCE_URL = "https://www.sgu.edu.cn/zsxxw/info/1022/12638.htm"
DEFAULT_INPUT = ROOT / "assets" / "raw-cache" / "sgu" / "2025-guangdong-admission.xlsx"
AGGREGATOR_SOURCE = "掌上高考（聚合补充）"

ADMISSION_COLUMNS = [
    "year",
    "province",
    "track",
    "batch",
    "school_name",
    "school_code",
    "major_group",
    "major_name",
    "plan_type",
    "min_score",
    "min_rank",
    "admit_count",
    "source_url",
    "source_name",
    "notes",
]


def require_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("缺少 openpyxl；请用 bundled Python runtime 运行本脚本。") from exc
    return load_workbook


def norm(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def numeric_text(value: object) -> str:
    text = norm(value).replace(",", "")
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return ""
    if number.is_integer():
        return str(int(number))
    return f"{number:.6f}".rstrip("0").rstrip(".")


def clean_major_name(value: object) -> str:
    text = norm(value)
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    return text


def infer_plan_type(major_name: str) -> str:
    if "学分互认" in major_name:
        return "学分互认"
    if "应用技术类" in major_name or "协同培养" in major_name:
        return "协同培养"
    return "普通类"


def track_from_batch(batch_context: str) -> str:
    if "历史" in batch_context:
        return "历史类"
    if "物理" in batch_context:
        return "物理类"
    return ""


def normalized_rows(args: argparse.Namespace) -> list[dict[str, str]]:
    load_workbook = require_openpyxl()
    workbook = load_workbook(args.input, data_only=True, read_only=True)
    worksheet = workbook["普通批"]
    rows: list[dict[str, str]] = []
    batch_context = ""
    current_group = ""

    for values in worksheet.iter_rows(min_row=2, max_col=10, values_only=True):
        raw_batch, raw_group, major_code, major_name_raw, subject_req, feature, _plan, admit_count, min_score, min_rank = values
        if raw_batch:
            batch_context = norm(raw_batch)
            current_group = ""
        if raw_group:
            current_group = norm(raw_group)

        if batch_context not in {"本科_普通类(历史科目组)", "本科_普通类(物理科目组)"}:
            continue
        track = track_from_batch(batch_context)
        if track != args.track:
            continue

        major_name = clean_major_name(major_name_raw)
        if not major_name or major_name == "小计" or "小计" in batch_context:
            continue
        score = numeric_text(min_score)
        rank = numeric_text(min_rank)
        count = numeric_text(admit_count)
        if not (current_group and score and rank and count):
            continue

        note_parts = [
            "学校官网附件《2025年录取统计（广东省内）.xlsx》",
            f"专业代码={norm(major_code)}" if norm(major_code) else "",
            f"选考科目={norm(subject_req)}" if norm(subject_req) else "",
            f"招生特征={norm(feature)}" if norm(feature) else "",
        ]
        rows.append(
            {
                "year": "2025",
                "province": args.province,
                "track": track,
                "batch": args.batch,
                "school_name": SCHOOL_NAME,
                "school_code": SCHOOL_CODE,
                "major_group": current_group,
                "major_name": major_name,
                "plan_type": infer_plan_type(major_name),
                "min_score": score,
                "min_rank": rank,
                "admit_count": count,
                "source_url": SOURCE_URL,
                "source_name": SOURCE_NAME,
                "notes": "；".join(part for part in note_parts if part),
            }
        )
    return rows


def read_existing(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_rows(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ADMISSION_COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def row_should_replace(row: dict[str, str], args: argparse.Namespace) -> bool:
    return (
        row.get("year") == "2025"
        and row.get("province") == args.province
        and row.get("track") == args.track
        and row.get("batch") == args.batch
        and row.get("school_name") == SCHOOL_NAME
        and row.get("source_name") in {SOURCE_NAME, AGGREGATOR_SOURCE}
    )


def import_rows(args: argparse.Namespace, rows: list[dict[str, str]]) -> tuple[int, int, int]:
    path = Path(args.data_dir).expanduser().resolve() / "admission_records.csv"
    existing = read_existing(path)
    before = len(existing)
    removed = 0
    if args.replace_existing:
        kept = []
        for row in existing:
            if row_should_replace(row, args):
                removed += 1
            else:
                kept.append(row)
        existing = kept
    write_rows(path, existing + rows)
    return before, removed, len(existing + rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Shaoguan University official 2025 Guangdong major scores.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--data-dir", required=True)
    parser.add_argument("--province", default="广东")
    parser.add_argument("--track", required=True, choices=["物理类", "历史类"])
    parser.add_argument("--batch", default="本科批")
    parser.add_argument("--replace-existing", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows = normalized_rows(args)
    by_plan = Counter(row["plan_type"] for row in rows)
    by_group = Counter(row["major_group"] for row in rows)
    print("# 韶关学院专业录取分数导入\n")
    print(f"- 来源：{SOURCE_URL}")
    print(f"- 输入：{Path(args.input).expanduser().resolve()}")
    print(f"- 范围：{args.province} / {args.track} / 2025")
    print(f"- 获取专业记录：{len(rows)}；类型：{dict(sorted(by_plan.items()))}；专业组数：{len(by_group)}")
    if rows:
        print(f"- 预览：{rows[:5]}")
    if args.dry_run:
        print("- 模式：预览，不写入")
        return
    before, removed, after = import_rows(args, rows)
    print(f"- 模式：写入；原记录 {before}，替换聚合/旧官网行 {removed}，写入后 {after}")


if __name__ == "__main__":
    main()
