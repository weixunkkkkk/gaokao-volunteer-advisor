#!/usr/bin/env python3
"""Normalize raw Gaokao CSV/XLSX tables into the advisor CSV schema."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Iterable


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATA_DIR = ROOT / "assets" / "data"

SCHEMA_COLUMNS = {
    "admission": [
        "year",
        "province",
        "track",
        "batch",
        "school_name",
        "school_code",
        "major_group",
        "major_name",
        "plan_type",
        "min_score",
        "min_rank",
        "admit_count",
        "source_url",
        "source_name",
        "notes",
    ],
    "rank": [
        "year",
        "province",
        "track",
        "score",
        "min_rank",
        "max_rank",
        "same_score_count",
        "source_url",
        "source_name",
    ],
    "majors": [
        "interest_keywords",
        "major_name",
        "degree_category",
        "employment_outlook",
        "typical_roles",
        "fit_notes",
        "risk_notes",
    ],
}

DEFAULT_OUTPUT = {
    "admission": "admission_records.csv",
    "rank": "rank_table.csv",
    "majors": "majors.csv",
}

ALIASES = {
    "year": ["year", "年份", "年度", "录取年份"],
    "province": ["province", "省份", "省市", "考生省份"],
    "track": ["track", "科类", "类别", "首选科目", "考试科类", "选科", "物理历史"],
    "batch": ["batch", "批次", "录取批次", "批次名称"],
    "school_name": ["school_name", "院校名称", "学校名称", "院校", "学校", "招生院校"],
    "school_code": ["school_code", "院校代码", "学校代码", "学校代号", "院校专业组代码", "院校代号"],
    "major_group": ["major_group", "专业组", "专业组代码", "院校专业组", "院校专业组号", "专业组代号"],
    "major_name": ["major_name", "专业名称", "专业", "专业类", "专业组名称", "专业备注"],
    "plan_type": ["plan_type", "计划类型", "招生类型", "类型", "类别", "投档类型"],
    "min_score": ["min_score", "最低分", "投档最低分", "录取最低分", "最低投档分", "分数线"],
    "min_rank": ["min_rank", "位次", "最低位次", "最低排位", "投档最低排位", "录取最低位次", "最低分位次"],
    "admit_count": ["admit_count", "录取人数", "计划数", "招生计划", "招生人数", "投档人数"],
    "source_url": ["source_url", "来源链接", "来源URL", "url", "URL"],
    "source_name": ["source_name", "来源名称", "来源", "发布单位"],
    "notes": ["notes", "备注", "说明", "特殊说明"],
    "score": ["score", "分数", "成绩", "总分", "文化总分"],
    "max_rank": ["max_rank", "最高排位", "累计人数", "本科累计人数", "累计位次", "本段最低位次", "最低排位"],
    "same_score_count": ["same_score_count", "同分人数", "本段人数", "人数", "分数段人数", "分数段人数（含本科加分）", "本科分数段人数", "本科分数段人数（含本科加分）"],
    "interest_keywords": ["interest_keywords", "兴趣关键词", "关键词", "方向关键词"],
    "degree_category": ["degree_category", "门类", "学科门类", "专业门类"],
    "employment_outlook": ["employment_outlook", "就业前景", "前景"],
    "typical_roles": ["typical_roles", "常见去向", "岗位", "就业方向"],
    "fit_notes": ["fit_notes", "适合人群", "匹配说明"],
    "risk_notes": ["risk_notes", "风险点", "注意事项"],
}


def norm(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized_header(value: str) -> str:
    return re.sub(r"\s+", "", norm(value).replace("\u3000", "")).lower()


def dedupe_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    output = []
    for index, header in enumerate(headers):
        base = norm(header) or f"column_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        if counts[base] == 1:
            output.append(base)
        else:
            output.append(f"{base}_{counts[base]}")
    return output


def read_csv_rows(path: Path, header_row: int | None) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        raw_rows = list(csv.reader(handle))
    return rows_from_matrix(raw_rows, header_row)


def read_xlsx_rows(path: Path, sheet_name: str | None, header_row: int | None) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "读取 Excel 需要 openpyxl。请用 Codex 内置 Python 运行，或先把 Excel 另存为 CSV。"
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"找不到工作表：{sheet_name}。可用工作表：{', '.join(workbook.sheetnames)}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]
    matrix = [[norm(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    return rows_from_matrix(matrix, header_row)


def rows_from_matrix(matrix: list[list[object]], header_row: int | None) -> tuple[list[str], list[dict[str, str]]]:
    if not matrix:
        return [], []
    if header_row is None:
        header_index = next((idx for idx, row in enumerate(matrix) if any(norm(cell) for cell in row)), 0)
    else:
        header_index = max(header_row - 1, 0)

    headers = dedupe_headers([norm(cell) for cell in matrix[header_index]])
    rows = []
    for raw in matrix[header_index + 1 :]:
        values = [norm(cell) for cell in raw]
        if not any(values):
            continue
        padded = values + [""] * max(0, len(headers) - len(values))
        rows.append({headers[idx]: padded[idx] for idx in range(len(headers))})
    return headers, rows


def read_input(path: Path, sheet_name: str | None, header_row: int | None) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path, sheet_name, header_row)
    if suffix == ".csv":
        return read_csv_rows(path, header_row)
    raise SystemExit(f"暂不支持的文件类型：{path.suffix}。请使用 CSV 或 XLSX。")


def parse_pairs(values: list[str], label: str) -> dict[str, str]:
    parsed = {}
    for item in values:
        if "=" not in item:
            raise SystemExit(f"{label} 参数必须是 key=value：{item}")
        key, value = item.split("=", 1)
        key = key.strip()
        if not key:
            raise SystemExit(f"{label} 参数 key 不能为空：{item}")
        parsed[key] = value.strip()
    return parsed


def detect_mapping(kind: str, headers: list[str], explicit_mapping: dict[str, str]) -> dict[str, str]:
    header_lookup = {normalized_header(header): header for header in headers if header}
    mapping = dict(explicit_mapping)
    for target in SCHEMA_COLUMNS[kind]:
        if target in mapping:
            continue
        for alias in ALIASES.get(target, []):
            source = header_lookup.get(normalized_header(alias))
            if source is not None:
                mapping[target] = source
                break
    return mapping


def clean_number(value: str) -> str:
    value = norm(value).replace(",", "").replace("，", "")
    if value.endswith(".0"):
        value = value[:-2]
    match = re.search(r"\d+(?:\.\d+)?", value)
    if match:
        value = match.group(0)
    return value


def split_prefixed_code(value: str) -> tuple[str, str] | None:
    value = re.sub(r"\s+", "", norm(value))
    match = re.match(r"^([A-Za-z0-9]{2,8})(.+)$", value)
    if not match:
        return None
    code, name = match.groups()
    if not name or name[0].isascii():
        return None
    return code, name


def split_group_suffix(value: str) -> tuple[str, str] | None:
    value = norm(value)
    match = re.match(r"^(.+?)[（(]([A-Za-z0-9]+)[）)]$", value)
    if not match:
        return None
    school_name, group = match.groups()
    return school_name.strip(), group.strip()


def apply_postprocess(
    row: dict[str, str],
    split_school_prefix: bool,
    strip_major_code_prefix: bool,
    split_major_group_suffix: bool = False,
) -> dict[str, str]:
    if split_school_prefix and row.get("school_name"):
        split = split_prefixed_code(row["school_name"])
        if split:
            code, name = split
            if not row.get("school_code") or row.get("school_code") == row.get("school_name"):
                row["school_code"] = code
            row["school_name"] = name
    if split_major_group_suffix and row.get("school_name"):
        split = split_group_suffix(row["school_name"])
        if split:
            school_name, major_group = split
            row["school_name"] = school_name
            if not row.get("major_group") or row.get("major_group") == row.get("school_name"):
                row["major_group"] = major_group
    if strip_major_code_prefix and row.get("major_name"):
        split = split_prefixed_code(row["major_name"])
        if split:
            _, name = split
            row["major_name"] = name
    return row


def normalize_row(
    kind: str,
    row: dict[str, str],
    mapping: dict[str, str],
    constants: dict[str, str],
) -> dict[str, str]:
    output = {}
    for column in SCHEMA_COLUMNS[kind]:
        if column in constants:
            value = constants[column]
        elif column in mapping:
            value = row.get(mapping[column], "")
        else:
            value = ""
        if column in {
            "year",
            "min_score",
            "min_rank",
            "admit_count",
            "score",
            "max_rank",
            "same_score_count",
        }:
            value = clean_number(value)
        output[column] = norm(value)
    if kind == "rank" and not output.get("min_rank") and output.get("max_rank") and output.get("same_score_count"):
        try:
            output["min_rank"] = str(int(output["max_rank"]) - int(output["same_score_count"]) + 1)
        except ValueError:
            pass
    return output


def has_content(row: dict[str, str], kind: str) -> bool:
    if kind == "admission":
        return bool(row.get("school_name") and (row.get("min_score") or row.get("min_rank")))
    if kind == "rank":
        return bool(row.get("score") and (row.get("min_rank") or row.get("max_rank")))
    if kind == "majors":
        return bool(row.get("major_name") and row.get("interest_keywords"))
    return any(row.values())


def output_path_for(kind: str, output_dir: Path, output_file: str | None) -> Path:
    if output_file:
        return Path(output_file).expanduser().resolve()
    return output_dir / DEFAULT_OUTPUT[kind]


def write_rows(path: Path, kind: str, rows: list[dict[str, str]], append: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = SCHEMA_COLUMNS[kind]
    mode = "a" if append and path.exists() else "w"
    write_header = mode == "w" or path.stat().st_size == 0
    with path.open(mode, encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        if write_header:
            writer.writeheader()
        writer.writerows(rows)


def preview_rows(rows: Iterable[dict[str, str]], limit: int) -> list[dict[str, str]]:
    return [row for idx, row in enumerate(rows) if idx < limit]


def build_constants(args: argparse.Namespace) -> dict[str, str]:
    constants = parse_pairs(args.const, "--const")
    quick = {
        "year": args.year,
        "province": args.province,
        "track": args.track,
        "batch": args.batch,
        "plan_type": args.plan_type,
        "source_url": args.source_url,
        "source_name": args.source_name,
        "notes": args.notes,
    }
    for key, value in quick.items():
        if value is not None:
            constants[key] = value
    return constants


def infer_plan_type(row: dict[str, str]) -> str | None:
    text = " ".join(row.values())
    if "中外合作" in text or "中外合办" in text:
        return "中外合作"
    if "联合培养" in text:
        return "联合培养"
    if "地方专项" in text:
        return "地方专项"
    if "少数民族" in text:
        return "少数民族班"
    if "预科" in text:
        return "预科班"
    return None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize raw Gaokao CSV/XLSX data into advisor CSV files.")
    parser.add_argument("--kind", choices=["admission", "rank", "majors"], required=True)
    parser.add_argument("--input", required=True, help="Raw CSV/XLSX file")
    parser.add_argument("--sheet", help="Excel sheet name; defaults to first sheet")
    parser.add_argument("--header-row", type=int, help="1-based header row number; defaults to first non-empty row")
    parser.add_argument("--data-dir", default=str(DEFAULT_DATA_DIR), help="Output data directory")
    parser.add_argument("--output-file", help="Override exact output CSV path")
    parser.add_argument("--mapping", action="append", default=[], help="target=source header mapping; repeatable")
    parser.add_argument("--const", action="append", default=[], help="target=value constant; repeatable")
    parser.add_argument("--year", help="Constant year")
    parser.add_argument("--province", help="Constant province")
    parser.add_argument("--track", help="Constant track")
    parser.add_argument("--batch", help="Constant batch")
    parser.add_argument("--plan-type", help="Constant plan_type")
    parser.add_argument("--infer-plan-type", action="store_true", help="Infer special plan_type from row text")
    parser.add_argument("--source-url", help="Constant source_url")
    parser.add_argument("--source-name", help="Constant source_name")
    parser.add_argument("--notes", help="Constant notes")
    parser.add_argument("--split-school-prefix", action="store_true", help="Split values like A001北京大学 into school_code=A001 and school_name=北京大学")
    parser.add_argument("--split-major-group-suffix", action="store_true", help="Split values like 复旦大学(01) into school_name=复旦大学 and major_group=01")
    parser.add_argument("--strip-major-code-prefix", action="store_true", help="Strip leading major code from values like 17文科试验班类")
    parser.add_argument("--append", action="store_true", help="Append to output CSV instead of overwriting")
    parser.add_argument("--dry-run", action="store_true", help="Print detected mapping and preview without writing")
    parser.add_argument("--format", choices=["markdown", "json"], default="markdown")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f"输入文件不存在：{input_path}")

    headers, raw_rows = read_input(input_path, args.sheet, args.header_row)
    explicit_mapping = parse_pairs(args.mapping, "--mapping")
    mapping = detect_mapping(args.kind, headers, explicit_mapping)
    constants = build_constants(args)
    normalized_rows = []
    for raw in raw_rows:
        row_constants = dict(constants)
        if args.infer_plan_type and args.kind == "admission":
            inferred = infer_plan_type(raw)
            if inferred:
                row_constants["plan_type"] = inferred
        row = normalize_row(args.kind, raw, mapping, row_constants)
        if args.kind == "admission":
            row = apply_postprocess(
                row,
                args.split_school_prefix,
                args.strip_major_code_prefix,
                args.split_major_group_suffix,
            )
        if has_content(row, args.kind):
            normalized_rows.append(row)
    output_path = output_path_for(args.kind, Path(args.data_dir).expanduser().resolve(), args.output_file)

    result = {
        "input": str(input_path),
        "kind": args.kind,
        "headers": headers,
        "mapping": mapping,
        "constants": constants,
        "raw_rows": len(raw_rows),
        "normalized_rows": len(normalized_rows),
        "output": str(output_path),
        "preview": preview_rows(normalized_rows, 5),
        "dry_run": args.dry_run,
    }

    if not args.dry_run:
        write_rows(output_path, args.kind, normalized_rows, args.append)

    if args.format == "json":
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(f"# 数据规范化导入\n")
        print(f"- 输入：{result['input']}")
        print(f"- 类型：{result['kind']}")
        print(f"- 原始行数：{result['raw_rows']}")
        print(f"- 有效导入行数：{result['normalized_rows']}")
        print(f"- 输出：{result['output']}")
        print(f"- 模式：{'预览，不写入' if args.dry_run else ('追加' if args.append else '覆盖')}")
        print("\n## 字段映射")
        for target in SCHEMA_COLUMNS[args.kind]:
            source = mapping.get(target)
            const = constants.get(target)
            if const is not None:
                print(f"- `{target}` = 常量 `{const}`")
            elif source:
                print(f"- `{target}` <- `{source}`")
            else:
                print(f"- `{target}` <- 空")
        print("\n## 预览")
        print(json.dumps(result["preview"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
